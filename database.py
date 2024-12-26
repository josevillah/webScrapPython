from PIL import Image
import os

from openpyxl import load_workbook

from connection import Connection

class ExcelReader:
    def __init__(self):
        self.workbook = None

    def load_file(self, file_name):
        """
        Carga el archivo Excel.
        """
        try:
            self.workbook = load_workbook(file_name)
            print(f"Archivo '{file_name}' cargado correctamente.")
        except FileNotFoundError:
            print(f"El archivo '{file_name}' no fue encontrado.")
        except Exception as e:
            print(f"Ocurrió un error al cargar el archivo: {e}")

    def read_sheet(self, sheet_name):
        """
        Lee los datos de una hoja específica.
        """
        if not self.workbook:
            print("El archivo no está cargado.")
            return None

        try:
            sheet = self.workbook[sheet_name]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            return data
        except KeyError:
            print(f"La hoja '{sheet_name}' no existe en el archivo.")
            return None
        except Exception as e:
            print(f"Error al leer la hoja: {e}")
            return None

    def copyImage(self, id, codpro, gender):
        # Determinar la subcarpeta según el género
        # subcat = "hombre" if gender == 'ROPA HOMBRE' else "mujer"

        # Definir rutas base y destino
        base_path = os.path.join(os.getcwd(), "registrar")
        destination_path = os.path.join("C:", os.sep, "xampp", "htdocs", "isaflor.cl", "public", "img", "productos")

        # Crear la carpeta de destino si no existe
        os.makedirs(destination_path, exist_ok=True)

        # Buscar la imagen que coincida con el nombre codpro
        file_name = f"{codpro}.jpg"  # Asume que la extensión es .jpg (puedes ajustar según tus necesidades)
        file_path = os.path.join(base_path, file_name)
        
        if os.path.isfile(file_path):
            print(f"Imagen guardada como WebP: {file_path}")
        else:
            print(f"No se encontró el archivo {file_name} en {base_path}")

        if os.path.isfile(file_path):
            try:
                # Abrir la imagen original
                with Image.open(file_path) as img:
                    # Transformar y guardar en formato WebP
                    webp_filename = f"{id}.webp"
                    webp_path = os.path.join(destination_path, webp_filename)
                    img.save(webp_path, format="WEBP", quality=85)  # Calidad ajustable (85 recomendado)
                    print(f"Imagen guardada como WebP: {webp_path}")
            except Exception as e:
                print(f"Error procesando el archivo {file_name}: {e}")
        else:
            print(f"No se encontró el archivo {file_name} en {base_path}")


    def insertData(self, data):
        connect = Connection()
        connect.connect()
        
        # Obtener el último ID en la tabla productos
        query = "SELECT id FROM productos ORDER BY id DESC LIMIT 1"
        result = connect.execute_query(query)
        id = (result[0][0] + 1) if result else 1  # Si no hay registros, comienza en 1
        
        if id:
            for row in data:
                verificar = "SELECT id FROM productos WHERE codpro = %s"

                marca = "LEE"
                
                query = """
                INSERT INTO productos (
                    id, codpro, nompro, anchpro, largpro, prepro, preoferpro, despro, marcapro, idsubcat, oculto, fecharegistro, urlimagen, medida, cantidad, agregarCarrito
                ) VALUES (%s, %s, %s, '', '', %s, '', '', %s, %s, %s, NOW(), %s, %s, %s, %s)
                """
                values = (
                    id,         # ID autoincrementado
                    row[0],     # Código del producto
                    row[1],     # Nombre del producto
                    row[2],     # Precio del producto
                    marca,
                    row[3],     # Subcategoría
                    0,     # Oculto
                    f"public/img/productos/{id}.webp",  # URL de imagen
                    0,       # Medida (ejemplo fija)
                    1,          # Cantidad inicial
                    0           # Agregar carrito (fijo en 0)
                )
                
                # Ejecutar la consulta de inserción
                try:
                    # Verificar si el producto ya existe
                    result = connect.execute_query(verificar, (row[0],))

                    if not result:
                        print(f"Insertando producto {row[0]}")
                        connect.connection.cursor().execute(query, values)
                        connect.connection.commit()  # Confirmar cambios en la base de datos
                        self.copyImage(id, row[0], row[3])
                    else:
                        print(f"Producto {row[0]} ya existe en la base de datos")

                    id += 1  # Incrementar ID para el próximo registro
                except Exception as e:
                    print(f"Error al insertar los datos: {e}")
                    connect.connection.rollback()  # Revertir cambios en caso de error

                # break
        
        connect.close()



# Instanciar el lector de Excel
reader = ExcelReader()

# Cargar el archivo de datos original
reader.load_file("./Reporte_productos_faltantes 2024-12-23 16_54_26.xlsx")
data = reader.read_sheet("Worksheet")

# Cargar el archivo de información combinada
reader.load_file("./Productos_combinados_2024-12-23_16-45-22.xlsx")
info = reader.read_sheet("Productos")

# Lista para guardar los resultados
arrayToInsert = []

if data:
    for row in data:
        if len(row) > 3 and (row[3] == 'ROPA HOMBRE' or row[3] == 'ROPA MUJER'):
            for row2 in info:
                if str(row[0]).strip() == str(row2[0]).strip():

                    if(row[3] == 'ROPA HOMBRE'):
                        idsubcat = 73
                    if(row[3] == 'ROPA MUJER'):
                        idsubcat = 74
                    
                    if(row[5] > 1):
                        stock = 1
                    else:
                        stock = 0

                    arrayToInsert.append(
                        [row[0], row2[1], row[4], idsubcat, stock, row[3]]
                    )


excel = ExcelReader()

excel.insertData(arrayToInsert)